# Copyright (c) 2024 NVIDIA CORPORATION & AFFILIATES.
# All rights reserved.
#
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Utilities to explore a TensorRT engine plan and its profiling data.

This module is a cookbook-friendly, `pandas`-free re-implementation of the core
of NVIDIA's `trt-engine-explorer` (a.k.a. `trex`).  TensorRT can export two
kinds of JSON files that describe a *built* engine:

  * a *graph* JSON  (``IEngineInspector`` output) - the final inference graph
    (layers, tensors, tactics, precisions, ...).
  * a *profiling* JSON  (``trtexec --exportProfile``) - per-layer latency.

Together with the *metadata* and *timing* JSON files produced by ``trtexec``,
these describe the engine well enough to analyse structure and performance
*without a GPU*.  This module loads them into plain Python objects and NumPy
arrays so the cookbook examples can compute statistics and draw figures with
Matplotlib.

The central object is :class:`EnginePlan`.  Instead of a `pandas.DataFrame`, the
per-layer table is exposed as ``plan.records`` (a ``list`` of ``dict``) and the
helper :meth:`EnginePlan.col` returns a NumPy array for any column.
"""

import json
import ntpath
import re
import warnings
from copy import deepcopy
from typing import Dict, List, Tuple

import numpy as np

__all__ = [
    "Activation",
    "Layer",
    "EnginePlan",
    "fold_no_ops",
    "read_graph_file",
    "read_profiling_file",
    "read_timing_file",
    "get_device_properties",
    "get_performance_summary",
    "get_builder_config",
    "import_graph_file",
    "summary_dict",
    "print_summary",
    "json_summary",
    "compute_precision_stats",
    "print_precision_stats",
    "json_precision_stats",
    "group_sum",
    "group_mean",
    "group_count",
    "precision_colormap",
    "layer_colormap",
    "colors_for",
    "clean_layer_name",
    "build_engine_graph",
    "render_engine_graph",
    "annotate_convolutions",
    "create_activations",
    "lint_convolutions",
    "lint_reformats",
    "lint_slices",
    "lint_qdq",
    "lint_engine",
    "parse_build_log",
    "parse_profiling_log",
    "write_build_metadata",
    "write_profiling_metadata",
    "export_engine_to_onnx",
    "write_engine_excel",
    "query_device_info",
    "sample_gpu_state",
    "get_max_clocks",
    "EngineArchive",
    "summarize_engine_tactics",
]

# ------------------------------------------------------------------------------
# Color maps (Matplotlib-compatible; ported from trex/colors.py)
# ------------------------------------------------------------------------------

_UNKNOWN_KEY_COLOR = "gray"

# A stable color per precision datatype (valid Matplotlib color specs).
precision_colormap = {
    "INT8": "#76b900",  # NVIDIA green
    "FP32": "red",
    "FP16": "orange",
    "INT32": "lightgray",
    "INT64": "silver",
    "FP8": "deepskyblue",
    "INT4": "#CC66FF",
    "FP4": "#F4D03F",
}

# A stable color per layer type.
layer_colormap = {
    "Convolution": "#4682B4",
    "Deconvolution": "#7B68EE",
    "ConvActPool": "#6495ED",
    "MatrixMultiply": "#1E90FF",
    "gemm": "#1E90FF",
    "Reformat": "#00FFFF",
    "Shuffle": "#BC8F8F",
    "Slice": "#FFA500",
    "Scale": "#8FBC8B",
    "Quantize": "#6B8E23",
    "Pooling": "#3CB371",
    "PluginV2": "#C71585",
    "PointWise": "#9ACD32",
    "ElementWise": "#9ACD32",
    "Reduce": "#90EE90",
    "SoftMax": "#DA70D6",
    "Myelin": "#B39C4D",
    "kgen": "#B39C4D",
    "NonZero": "#98FB98",
    "TrainStation": "#FFA07A",
}

def colors_for(keys, colormap: Dict) -> List[str]:
    """Map an iterable of keys to colors, falling back to a neutral gray."""
    return [colormap.get(k, _UNKNOWN_KEY_COLOR) for k in keys]

# ------------------------------------------------------------------------------
# Activation (tensor region) abstraction
# ------------------------------------------------------------------------------

# Compress JSON's long "Format/Datatype" description strings into short labels.
_region_format_dict = {
    "Four wide channel vectorized row major Int8 format": "Int8 NC/4HW4",
    "Four wide channel vectorized row major FP32 format": "FP32 NC/4HW4",
    "Thirty-two wide channel vectorized row major Int8 format": "Int8 NC/32HW32",
    "Thirty-two wide channel vectorized row major FP32 format": "FP32 NC/32HW32",
    "Thirty-two wide channel vectorized row major FP16 format": "FP16 NC/32HW32",
    "Thirty-two wide channel vectorized row major Int8 format with 3 spatial dimensions": "Int8 NC32DHW",
    "Thirty-two wide channel vectorized row major FP16 format with 3 spatial dimensions": "FP16 NC32DHW",
    "Sixteen wide channel vectorized row major FP16 format": "FP16 NC16HW",
    "Channel major FP16 format where channel % 4 == 0": "FP16 NHWC4",
    "Channel major FP32 format where channel % 4 == 0": "FP32 NHWC4",
    "Channel major Int8 format where channel % 4 == 0": "Int8 NHWC4",
    "Channel major FP16 format where channel % 8 == 0": "FP16 NHWC8",
    "Channel major FP16 format where channel % 16 == 0": "FP16 NHWC16",
    "Channel major FP16 format where channel == 4 and column stride % 32 == 0": "FP16 NHWC4",
    "Channel major INT8 format where channel == 4 and column stride % 32 == 0": "Int8 NHWC4",
    "Channel major FP16 format where channel % 2 == 0": "FP16 NHWC2",
    "Channel major INT8 format where column stride % 32 == 0": "Int8 NHWC1",
    "Channel major INT8 format where channel % 16 == 0": "Int8 NHWC16",
    "Row major INT8 format where column stride % 64 == 0": "Int8 NCHW",
    "Channel major FP16 format where channel % 8 == 0 with 3 spatial dimensions": "FP16 NDHWC8",
    "Channel major FP16 format where channel == 1 and column stride % 32 == 0": "FP16 NHWC1",
    "Row major FP16 format where column stride % 64 == 0": "FP16",
    "Two wide channel vectorized row major FP16 format": "FP16 NC/2HW2",
    "Row major linear FP32": "FP32 NCHW",
    "Row major linear Int32": "INT32 NCHW",
    "Row major linear FP16 format": "FP16 NCHW",
    "Row major Int8 format": "Int8 NCHW",
    "Channel major FP32 format": "FP32 NHWC",
    "Channel major FP16 format": "FP16 NHWC",
    "Channel major Int8 format": "Int8 NHWC",
    "Row major linear BOOL": "Bool",
    "Channel major FP32 format with 3 spatial dimensions": "FP32 NDHWC",
    "Channel major FP32 format with 3 spatial dimensions where channel % 4 == 0": "FP32 NDHWC4",
    "Channel major FP32 format where channel % 4 == 0 with 3 spatial dimensions": "FP32 NDHWC4",
    "Row major linear UInt8 format": "UInt8 NCHW",
    "Channel major UInt8 format": "UInt8 NHWC",
    "Row major linear Int64 format": "Int64 NCHW",
    "Row major linear BFloat16 format": "BF16 NCHW",
    "Channel major BFloat16 format where channel % 8 == 0": "BF16 NHWC8",
    "Channel major BFloat16 format where channel % 4 == 0": "BF16 NHWC4",
    "Channel major BFloat16 format where channel % 8 == 0 with 3 spatial dimensions": "BF16 NDHWC8",
    "Channel major BFloat16 format where channel % 2 == 0": "BF16 NHWC2",
    "Two wide channel vectorized row major BFloat16 format": "BF16 NC2HW",
    "Row major linear FP4 format (kLINEAR)": "FP4",
    "Row major linear FP4 format": "FP4",
    "Row major linear FP8 format": "FP8 NCHW",
    "Unknown format": "Unknown format",
    # kgen formats
    "BFloat16": "BFloat16",
    "Bool": "Bool",
    "Double": "Double",
    "DoubleComplex": "DoubleComplex",
    "Float": "Float",
    "FloatComplex": "FloatComplex",
    "FP8": "FP8",
    "Half": "Half",
    "Int16": "Int16",
    "Int32": "Int32",
    "Int64": "Int64",
    "Int8": "Int8",
    "None": "None",
    "UInt16": "UInt16",
    "UInt32": "UInt32",
    "UInt64": "UInt64",
    "Int4": "Int4",
    "FP4": "FP4",
    "FP4E2M1": "FP4",
}

# Map a short data-type label to (precision, bytes-per-element).
_precision_dict = {
    "FP8": ("FP8", 1),
    "FP16": ("FP16", 2),
    "Half": ("FP16", 2),
    "FP32": ("FP32", 4),
    "Float": ("FP32", 4),
    "Double": ("FP64", 8),
    "BFloat16": ("FP32", 2),
    "Int8": ("INT8", 1),
    "Int16": ("INT16", 2),
    "INT32": ("INT32", 4),
    "Int32": ("INT32", 4),
    "Int64": ("INT64", 8),
    "UInt8": ("UINT8", 1),
    "UInt16": ("UINT16", 2),
    "UInt32": ("UINT32", 4),
    "UInt64": ("UINT64", 8),
    "Int4": ("INT4", 0.5),
    "FP4": ("FP4", 0.5),
}

class Activation:
    """Convenience wrapper around a plan tensor ("region") described in JSON."""

    def __init__(self, raw_dict: Dict):
        self.name = raw_dict["Name"]
        self.shape = raw_dict["Dimensions"]
        fmt = raw_dict["Format/Datatype"].replace(".", "")
        self.format = _region_format_dict.get(fmt, "Unknown format")
        data_type = self.format.split(" ")[0]
        self.precision, self.data_size = _precision_dict.get(data_type, (data_type, 0))
        self.size_bytes = int(np.prod(self.shape) * self.data_size)

    def tooltip(self) -> str:
        return "\\n".join((str(self.shape), self.format))

    def __repr__(self) -> str:
        return f"{self.name}: {str(self.shape)}x{self.format}"

def create_activations(layer: Dict) -> Tuple[List[Activation], List[Activation]]:
    """Build the input/output :class:`Activation` lists of a raw layer dict."""
    inputs = [Activation(tensor) for tensor in layer["Inputs"]]
    outputs = [Activation(tensor) for tensor in layer["Outputs"]]
    return inputs, outputs

# ------------------------------------------------------------------------------
# Layer abstraction
# ------------------------------------------------------------------------------

# Bytes per element for weight/bias constant blobs.
_constant_size_dict = {"Int8": 1, "Half": 2, "Float": 4, "Int32": 4, "FP16": 2, "FP32": 4}

class Layer:
    """A single engine-plan layer, parsed from its raw JSON dict."""

    def __init__(self, raw_dict: Dict):
        self.raw_dict = raw_dict
        self.name = raw_dict["Name"]
        self.type = raw_dict.get("ParameterType", raw_dict["LayerType"])
        self.metadata = raw_dict.get("Metadata", None)
        self.subtype = raw_dict["LayerType"]
        self.inputs = [Activation(t) for t in raw_dict["Inputs"]]
        self.outputs = [Activation(t) for t in raw_dict["Outputs"]]
        self.outputs_size_bytes = int(np.sum([o.size_bytes for o in self.outputs]))
        if self.inputs:
            self.precision = self.inputs[0].precision
            self.inputs_size_bytes = int(np.sum([i.size_bytes for i in self.inputs]))
        else:
            self.precision = None
            self.inputs_size_bytes = 0
        self.total_io_size_bytes = self.inputs_size_bytes + self.outputs_size_bytes
        self._parse_weights()
        self.total_footprint_bytes = self.total_io_size_bytes + self.weights_size

    @staticmethod
    def _parse_constant(const: Dict) -> Tuple[int, str, int]:
        cnt = const["Count"]
        data_type = const["Type"]
        data_size = _constant_size_dict[data_type]
        return cnt, data_type, cnt * data_size

    def _parse_weights(self):
        try:
            self.weights_cnt, self.weights_type, self.weights_size = Layer._parse_constant(self.raw_dict["Weights"])
        except KeyError:
            self.weights_cnt, self.weights_type, self.weights_size = 0, None, 0
        try:
            self.bias_cnt, self.bias_type, self.bias_size = Layer._parse_constant(self.raw_dict["Bias"])
        except KeyError:
            self.bias_cnt, self.bias_type, self.bias_size = 0, None, 0

    def tooltip(self) -> str:
        skip = ("InputRegions", "OutputRegions", "Inputs", "Outputs", "ParameterType", "LayerName")
        return "".join(f"{k}:{v}\\n" for k, v in sorted(self.raw_dict.items()) if k not in skip)

    def __repr__(self) -> str:
        return f"Layer({self.name})"

def fold_no_ops(layers: List[Layer], bindings: List[str]) -> List[Layer]:
    """Remove ``NoOp`` layers, rewiring their producers/consumers around them."""

    def consumers_producers_dict(layers):
        consumers, producers = {}, {}
        for layer in layers:
            for loc, inp in enumerate(layer.inputs):
                consumers.setdefault(inp.name, []).append((layer.name, loc))
            for loc, outp in enumerate(layer.outputs):
                producers.setdefault(outp.name, []).append((layer.name, loc))
        return consumers, producers

    def fold(no_op: Layer):
        try:
            for successor_name, in_port in activation_consumers[no_op.outputs[0].name]:
                # NoOp has a single input, so port #0 is safe.
                ret[successor_name].inputs[in_port] = no_op.inputs[0]
        except KeyError:
            # A leaf NoOp whose output is a binding: move it back to the producer.
            if no_op.outputs[0].name in bindings:
                for predecessor_name, out_port in activation_producers[no_op.inputs[0].name]:
                    ret[predecessor_name].outputs[out_port] = no_op.outputs[0]

    ret = {layer.name: layer for layer in layers}
    activation_consumers, activation_producers = consumers_producers_dict(layers)
    for layer in layers:
        if layer.type == "NoOp":
            fold(layer)
    return [layer for layer in ret.values() if layer.type != "NoOp"]

# ------------------------------------------------------------------------------
# JSON parsing
# ------------------------------------------------------------------------------

def _read_json(json_file):
    try:
        return json.load(json_file)
    except Exception:
        raise ValueError(f"Could not load JSON file {json_file}")

def read_graph_file(graph_file: str) -> Tuple[List, List]:
    """Read a graph JSON file, returning ``(raw_layers, bindings)``."""
    err_msg = f"File {graph_file} does not conform to the expected JSON format."
    with open(graph_file) as f:
        graph = _read_json(f)
    if not isinstance(graph, dict):
        raise ValueError(err_msg)
    layers = graph["Layers"]
    bindings = graph.get("Bindings", [])  # Older TRT did not include bindings.
    if not isinstance(layers, list):
        raise ValueError(err_msg)
    if not isinstance(layers[0], dict):
        raise ValueError(err_msg + "\nMake sure to enable detailed ProfilingVerbosity when building the engine.")
    return layers, bindings

def read_profiling_file(profiling_file: str) -> List[Dict]:
    """Read a profiling JSON file, dropping the leading summary/count record."""
    with open(profiling_file) as f:
        perf = _read_json(f)
    return [rec for rec in perf if len(rec) in (4, 5)]

def read_metadata_file(metadata_file: str, device: int = 0):
    with open(metadata_file) as f:
        return _read_json(f)[device]

def read_timing_file(timing_json_file: str) -> List[float]:
    """Read a timing JSON file, returning the list of per-iteration latencies (ms)."""
    with open(timing_json_file) as f:
        recs = _read_json(f)
    return [rec["latencyMs"] for rec in recs]

def _read_perf_metadata_file(metadata_file: str, section: str):
    with open(metadata_file) as f:
        return _read_json(f)[section]

def get_device_properties(metadata_file: str) -> Dict:
    try:
        return _read_perf_metadata_file(metadata_file, "device_information")
    except (FileNotFoundError, TypeError):
        return {}

def get_performance_summary(metadata_file: str) -> Dict:
    try:
        return _read_perf_metadata_file(metadata_file, "performance_summary")
    except (FileNotFoundError, TypeError):
        return {}

def get_builder_config(metadata_file: str) -> Dict:
    try:
        d = _read_perf_metadata_file(metadata_file, "model_options")
        d.update(_read_perf_metadata_file(metadata_file, "build_options"))
        return d
    except (FileNotFoundError, TypeError):
        return {}

def import_graph_file(graph_file: str, profile_id: int = None) -> Tuple[List, List]:
    """Read and normalize a graph JSON file (fix names, dtypes, shape profiles)."""

    def filter_profiles(raw_layers, bindings, profile_id):
        """Keep only the layers/bindings that belong to one shape-profile.

        Names of layers that belong to profile N (N > 0) carry a ``[profile N]``
        suffix; the first profile has no suffix.
        """

        def use_name(name):
            belongs_to_some_profile = re.search(r"\[profile +[0-9]\]", name)
            if belongs_to_some_profile and not profile_id:
                return False
            if profile_id:
                if not re.search(rf"\[profile {profile_id}\]", name):
                    return False
            return True

        filtered_layers = [l for l in raw_layers if use_name(l["Name"])]
        filtered_bindings = [b for b in bindings if use_name(b)]
        if not filtered_layers or not filtered_bindings:
            raise ValueError(f"Something went wrong when filtering layers from profile ({profile_id}).\n"
                             "Most likely the profile data does not exist in the graph file, so try "
                             "without providing a profile id.")
        return filtered_layers, filtered_bindings

    def disambiguate_layer_names(raw_layers):
        """Append a numeric suffix to duplicated layer names."""
        names_cnt = {}
        for raw_layer in raw_layers:
            name = raw_layer["Name"]
            if name in names_cnt:
                names_cnt[name] += 1
                raw_layer["Name"] = name + "_" + str(names_cnt[name])
            else:
                names_cnt[name] = 1
        return raw_layers

    def convert_deconv(raw_layers):
        for raw_layer in raw_layers:
            if raw_layer.get("ParameterType") == "Convolution" and raw_layer.get("LayerType") == "CaskDeconvolutionV2":
                raw_layer["ParameterType"] = "Deconvolution"
        return raw_layers

    def unify_conv_name(raw_layers):
        """TRT 10.x renamed some convolutions to 'correlation'."""
        for raw_layer in raw_layers:
            if raw_layer.get("LayerType") == "correlation":
                raw_layer["ParameterType"] = "Convolution"
        return raw_layers

    def fix_unicode(raw_layers):
        """Replace non-ASCII separators that break SVG rendering (TRT 8.6 / 10.0)."""
        unit_sep, rec_sep, trex_sep = "\x1E", "\x1F", "+"
        replace = lambda s: s.replace(unit_sep, trex_sep).replace(rec_sep, trex_sep)
        for l in raw_layers:
            if "Name" in l:
                l["Name"] = replace(l["Name"])
            if "Metadata" in l:
                l["Metadata"] = replace(l["Metadata"])
        return raw_layers

    def remove_signal_wait(raw_layers):
        return [l for l in raw_layers if l["LayerType"] not in ("signal", "wait")]

    raw_layers, bindings = read_graph_file(graph_file)
    raw_layers = fix_unicode(raw_layers)
    raw_layers = convert_deconv(raw_layers)
    raw_layers = unify_conv_name(raw_layers)
    raw_layers = remove_signal_wait(raw_layers)
    raw_layers = disambiguate_layer_names(raw_layers)
    raw_layers, bindings = filter_profiles(raw_layers, bindings, profile_id)
    return raw_layers, bindings

# ------------------------------------------------------------------------------
# EnginePlan
# ------------------------------------------------------------------------------

# Latency columns added to each record. Names match the original trex JSON keys.
_LATENCY_COLS = ("latency.pct_time", "latency.avg_time", "latency.median_time", "latency.time")
_PROFILE_KEY_MAP = {
    "percentage": "latency.pct_time",
    "averageMs": "latency.avg_time",
    "medianMs": "latency.median_time",
    "timeMs": "latency.time",
}
# Standardize weight/bias type spellings onto the activation-precision vocabulary.
_WEIGHT_PRECISION_MAP = {"Int8": "INT8", "Half": "FP16", "Float": "FP32", "Int32": "INT32", "FP16": "FP16", "FP32": "FP32"}

class EnginePlan:
    """A parsed TensorRT engine plan and (optionally) its profiling data.

    The per-layer table lives in ``self.records`` (a ``list`` of ``dict``).  Each
    record merges the raw graph fields with a few computed columns
    (``type``, ``subtype``, ``tactic``, ``precision``, ``output_precision``,
    ``total_io_size_bytes``, ``weights_size``, ``total_footprint_bytes``) and,
    when a profiling file is supplied, the ``latency.*`` columns.

    Use :meth:`col` to pull any column out as a NumPy array.
    """

    def __init__(
        self,
        graph_file,
        profiling_file: str = None,
        profiling_metadata_file: str = None,
        build_metadata_file: str = None,
        name: str = None,
        profile_id: int = None,
    ):
        # 1. Load and normalize the engine graph.
        if isinstance(graph_file, str):
            self.name = name or _path_leaf(graph_file)
            raw_layers, self.bindings = import_graph_file(graph_file, profile_id)
        elif isinstance(graph_file, list):
            # Allow constructing directly from an in-memory list of raw layers.
            self.name = name or ""
            raw_layers, self.bindings = graph_file, [""]
        else:
            raise TypeError("graph_file must be a path (str) or a list of raw layers")
        raw_layers = [l for l in raw_layers if l["LayerType"] != "shape_call"]

        # 2. Build Layer objects, fold NoOps and separate out Constants.
        layers = [Layer(l) for l in raw_layers]
        layers = fold_no_ops(layers, self.bindings)
        self.all_layers = deepcopy(layers)
        self.constants = [l for l in layers if l.type == "Constant"]
        self.layers = [l for l in layers if l.type != "Constant"]

        # 3. Read the profiling data (skipping ignored layers).
        ignore = {l.name for l in layers if l.type in ("Constant", "NoOp")}
        self._raw_perf = None
        if profiling_file:
            self._raw_perf = [rec for rec in read_profiling_file(profiling_file) if rec["name"] not in ignore]

        # 4. Build the per-layer records table and merge profiling data.
        self.records = self._build_records()
        self._merge_profiling_data()

        # 5. Summary numbers and metadata.
        self.total_act_size = sum(l.total_io_size_bytes for l in self.layers)
        self.total_weights_size = sum(l.weights_size for l in self.layers)
        self.total_runtime = float(np.sum(self.col("latency.avg_time")))
        self.device_properties = get_device_properties(profiling_metadata_file)
        self.performance_summary = get_performance_summary(profiling_metadata_file)
        self.builder_cfg = get_builder_config(build_metadata_file)

    # -- construction helpers --------------------------------------------------

    def _build_records(self) -> List[Dict]:
        """One record per non-Constant/NoOp layer (aligned with ``self.layers``)."""
        records = []
        for layer in self.layers:
            rec = dict(layer.raw_dict)  # shallow copy of the raw graph fields
            rec["subtype"] = layer.subtype
            rec["type"] = layer.type
            rec["tactic"] = rec.get("TacticName", "TensorRT") or "TensorRT"
            rec["precision"] = layer.precision
            rec["total_io_size_bytes"] = layer.total_io_size_bytes
            rec["weights_size"] = layer.weights_size
            rec["total_footprint_bytes"] = layer.total_footprint_bytes
            rec["output_precision"] = layer.outputs[0].precision if layer.outputs else ""
            records.append(rec)
        return records

    def _merge_profiling_data(self):
        """Attach ``latency.*`` columns to every record."""
        if self._raw_perf is None:
            warnings.warn("Profiling data was not provided.")
            for rec in self.records:
                for col in _LATENCY_COLS:
                    rec[col] = 0
            return
        perf = [{_PROFILE_KEY_MAP.get(k, k): v for k, v in r.items()} for r in self._raw_perf]
        if len(perf) == len(self.records):
            for rec, p in zip(self.records, perf):
                for col in _LATENCY_COLS:
                    rec[col] = p.get(col, 0)
        else:
            warnings.warn(f"Partial profiling data: the number of layers in the engine graph "
                          f"({len(self.records)}) does not match the number of profiled layers "
                          f"({len(perf)}).\nThis can happen if you are not using the first shape-profile.")
            by_name = {p["name"]: p for p in perf}
            for rec in self.records:
                p = by_name.get(rec["Name"], {})
                for col in _LATENCY_COLS:
                    rec[col] = p.get(col, 0)

    # -- accessors -------------------------------------------------------------

    def col(self, name: str, records: List[Dict] = None) -> np.ndarray:
        """Return column ``name`` across ``records`` (default: all) as an ndarray."""
        records = self.records if records is None else records
        return np.array([rec.get(name) for rec in records], dtype=object)

    def get_layers_by_type(self, layer_type: str) -> List[Dict]:
        """Return the records whose ``type`` equals ``layer_type``."""
        return [rec for rec in self.records if rec["type"] == layer_type]

    def find(self, layer_name: str):
        for l in self.layers:
            if l.name == layer_name:
                return l
        return None

    def get_bindings(self) -> Tuple[List[Activation], List[Activation]]:
        """Return ``(input_bindings, output_bindings)`` as :class:`Activation` lists."""
        inputs, outputs, seen = [], [], set()
        for layer in self.layers:
            for inp in layer.inputs:
                if inp.name in self.bindings and inp.name not in seen:
                    inputs.append(inp)
                seen.add(inp.name)
            for outp in layer.outputs:
                if outp.name in self.bindings and outp.name not in seen:
                    outputs.append(outp)
                seen.add(outp.name)
        return list(set(inputs)), list(set(outputs))

    def summary(self):
        return print_summary(self)

    def precision_stats(self):
        return compute_precision_stats(self)

def _path_leaf(path: str) -> str:
    head, tail = ntpath.split(path)
    return tail or ntpath.basename(head)

# ------------------------------------------------------------------------------
# Grouping / aggregation helpers (pandas.groupby replacements)
# ------------------------------------------------------------------------------

def group_count(records: List[Dict], key: str) -> Dict[str, int]:
    """Count records per distinct value of ``key``."""
    out = {}
    for rec in records:
        out[rec[key]] = out.get(rec[key], 0) + 1
    return out

def group_sum(records: List[Dict], key: str, attr: str) -> Dict[str, float]:
    """Sum ``attr`` per distinct value of ``key``."""
    out = {}
    for rec in records:
        out[rec[key]] = out.get(rec[key], 0) + rec.get(attr, 0)
    return out

def group_mean(records: List[Dict], key: str, attr: str) -> Dict[str, float]:
    """Mean of ``attr`` per distinct value of ``key``."""
    total, count = {}, {}
    for rec in records:
        total[rec[key]] = total.get(rec[key], 0) + rec.get(attr, 0)
        count[rec[key]] = count.get(rec[key], 0) + 1
    return {k: total[k] / count[k] for k in total}

# ------------------------------------------------------------------------------
# Textual / JSON summaries
# ------------------------------------------------------------------------------

def summary_dict(plan: EnginePlan) -> Dict:
    """Create a dictionary of the most important attributes of an engine plan."""
    mb = 1024 * 1024
    bindings = plan.get_bindings()
    nl = "\n\t\t"
    return {
        "Inputs": nl.join(str(b) for b in bindings[0]),
        "Outputs": nl.join(str(b) for b in bindings[1]),
        "Average time": f"{plan.total_runtime:.3f} ms",
        "Layers": f"{len(plan.records)}",
        "Weights": f"{plan.total_weights_size / mb:.1f} MB",
        "Activations": f"{plan.total_act_size / mb:.1f} MB",
    }

def print_summary(plan: EnginePlan):

    def print_dict(d):
        for k, v in d.items():
            print(f"\t{k}: {v}")

    print("Model:")
    print_dict(summary_dict(plan))
    print("Device Properties:")
    print_dict(plan.device_properties)
    print("Builder Configuration:")
    print_dict(plan.builder_cfg)
    print("Performance Summary:")
    print_dict(plan.performance_summary)

def json_summary(plan: EnginePlan) -> str:
    return json.dumps(
        {
            "Model": summary_dict(plan),
            "Device Properties": plan.device_properties,
            "Builder Configuration": plan.builder_cfg,
            "Performance Summary": plan.performance_summary,
        },
        indent=2,
    )

# ------------------------------------------------------------------------------
# Precision statistics
# ------------------------------------------------------------------------------

def compute_precision_stats(plan: EnginePlan) -> Dict:
    """Total bytes of input activations, output activations and weights per precision."""
    input_activations, output_activations, weights = {}, {}, {}
    for layer in plan.layers:
        for act in layer.inputs:
            input_activations[act.precision] = input_activations.get(act.precision, 0) + act.size_bytes
        for act in layer.outputs:
            output_activations[act.precision] = output_activations.get(act.precision, 0) + act.size_bytes
        if layer.weights_size > 0 and layer.weights_type is not None:
            p = _WEIGHT_PRECISION_MAP.get(layer.weights_type, layer.weights_type)
            weights[p] = weights.get(p, 0) + layer.weights_size
        if layer.bias_size > 0 and layer.bias_type is not None:
            p = _WEIGHT_PRECISION_MAP.get(layer.bias_type, layer.bias_type)
            weights[p] = weights.get(p, 0) + layer.bias_size
    return {"input_activations": input_activations, "output_activations": output_activations, "weights": weights}

def _bytes_to_human_readable(size_bytes: float) -> str:
    for unit, limit in (("B", 1024), ("KB", 1024 ** 2), ("MB", 1024 ** 3)):
        if size_bytes < limit:
            return f"{size_bytes / (limit / 1024):.2f} {unit}" if unit != "B" else f"{size_bytes} B"
    return f"{size_bytes / 1024**3:.2f} GB"

def print_precision_stats(plan: EnginePlan):
    stats = compute_precision_stats(plan)

    def print_section(title, data):
        print(f"\n{title}:")
        print("-" * len(title))
        if not data:
            print("None")
            return
        total = sum(data.values())
        print(f"{'Precision':<10} {'Size':<12} {'Percentage':<10}")
        print("-" * 32)
        for precision, size in sorted(data.items(), key=lambda x: x[1], reverse=True):
            pct = (size / total * 100) if total > 0 else 0
            print(f"{precision:<10} {_bytes_to_human_readable(size):<12} {pct:.2f}%")
        print("-" * 32)
        print(f"{'Total':<10} {_bytes_to_human_readable(total):<12} 100.00%")

    print("\n=== Precision Statistics ===")
    print_section("Input Activations", stats["input_activations"])
    print_section("Output Activations", stats["output_activations"])
    print_section("Weights", stats["weights"])

def json_precision_stats(plan: EnginePlan) -> str:
    stats = compute_precision_stats(plan)
    result = {}
    for category, data in stats.items():
        total = int(sum(data.values()))
        result[category] = {"total_bytes": total, "by_precision": {}}
        for precision, size in data.items():
            size = int(size)
            pct = (size / total * 100) if total > 0 else 0
            result[category]["by_precision"][precision] = {"bytes": size, "percentage": pct}
    return json.dumps(result, indent=2)

# ------------------------------------------------------------------------------
# Engine-graph rendering (Graphviz)
# ------------------------------------------------------------------------------
#
# A compact re-implementation of trex/graphing.py's DotGraph: draw the engine as
# a directed graph with one node per layer (colored by layer type) and one edge
# per data dependency (colored by tensor precision). Graphviz is imported lazily
# so the rest of this module stays dependency-light and GPU-free.

def clean_layer_name(name: str) -> str:
    """Escape characters that confuse Graphviz record/HTML labels."""
    return name.replace("||", r"\|\|").replace("{", "").replace("}", "")

def _dot_id(name: str) -> str:
    """A stable, Graphviz-safe node id derived from a (possibly long) name."""
    return "n" + str(abs(hash(name)))

def build_engine_graph(
    plan: "EnginePlan",
    display_layer_names: bool = True,
    display_latency: bool = True,
    latency_type: str = "latency.avg_time",
    display_edge_details: bool = True,
    display_constants: bool = False,
    display_bindings: bool = True,
    highlight_layers: List[str] = None,
    max_name_len: int = 48,
):
    """Build a Graphviz ``Digraph`` for an engine plan.

    Nodes are layers (filled with :data:`layer_colormap`), edges are the tensors
    that connect a producer layer to its consumer layers (colored with
    :data:`precision_colormap`). Optionally draws the engine's input/output
    bindings as gray terminal nodes.

    Returns a ``graphviz.Digraph``; use :func:`render_engine_graph` to write a file.
    """
    import graphviz  # lazy: only needed for graph rendering

    highlight_layers = set(highlight_layers or [])
    layers = list(plan.layers)
    if display_constants:
        layers = layers + list(plan.constants)

    def short(name: str) -> str:
        return name if len(name) <= max_name_len else "..." + name[-(max_name_len - 3):]

    # Map every tensor name to its producer layer (from layer outputs).
    producer = {}
    for layer in layers:
        for outp in layer.outputs:
            producer[outp.name] = layer

    dot = graphviz.Digraph(name=plan.name or "engine")
    dot.attr("node", shape="rectangle", style="filled", fontname="Helvetica", fontsize="10")
    dot.attr("edge", fontname="Helvetica", fontsize="8")

    # Layer nodes.
    latency_by_name = {rec["Name"]: float(rec.get(latency_type, 0) or 0) for rec in plan.records}
    for layer in layers:
        label_lines = []
        if display_layer_names:
            label_lines.append(clean_layer_name(short(layer.name)))
        label_lines.append(layer.type)
        if display_latency:
            lat = latency_by_name.get(layer.name, 0)
            if lat:
                label_lines.append(f"{lat:.4f} ms")
        node_kwargs = {"fillcolor": layer_colormap.get(layer.type, _UNKNOWN_KEY_COLOR)}
        if layer.name in highlight_layers:
            node_kwargs.update(penwidth="4", color="red")
        dot.node(_dot_id(layer.name), "\n".join(label_lines), **node_kwargs)

    # Edges: producer layer -> consumer layer, one per consumed tensor.
    known = {layer.name for layer in layers}
    for layer in layers:
        for inp in layer.inputs:
            prod = producer.get(inp.name)
            label = inp.tooltip().replace("\\n", "\n") if display_edge_details else ""
            color = precision_colormap.get(inp.precision, _UNKNOWN_KEY_COLOR)
            if prod is not None and prod.name in known and prod.name != layer.name:
                dot.edge(_dot_id(prod.name), _dot_id(layer.name), label=label, color=color)
            elif display_bindings and inp.name in plan.bindings:
                # A graph input binding: draw a terminal source node.
                bid = "in_" + _dot_id(inp.name)
                dot.node(bid, inp.name, shape="oval", fillcolor="lightgray")
                dot.edge(bid, _dot_id(layer.name), label=label, color=color)

    # Graph output bindings: draw a terminal sink node per output binding.
    if display_bindings:
        for layer in layers:
            for outp in layer.outputs:
                if outp.name in plan.bindings:
                    bid = "out_" + _dot_id(outp.name)
                    dot.node(bid, outp.name, shape="oval", fillcolor="lightgray")
                    color = precision_colormap.get(outp.precision, _UNKNOWN_KEY_COLOR)
                    label = outp.tooltip().replace("\\n", "\n") if display_edge_details else ""
                    dot.edge(_dot_id(layer.name), bid, label=label, color=color)

    return dot

def render_engine_graph(plan: "EnginePlan", output_path: str, output_format: str = "svg", **kwargs) -> str:
    """Build and render the engine graph to ``output_path`` (extension stripped).

    ``output_format`` is any Graphviz output (``svg``, ``png``, ``dot``, ...).
    Returns the path of the written file.
    """
    import os

    dot = build_engine_graph(plan, **kwargs)
    dot.format = output_format
    base = os.path.splitext(str(output_path))[0]
    out = dot.render(outfile=f"{base}.{output_format}", view=False, overwrite_source=False)
    return out

# ------------------------------------------------------------------------------
# Convolution analysis (implicit GEMM)
# ------------------------------------------------------------------------------

def annotate_convolutions(plan: "EnginePlan") -> List[Dict]:
    """Enrich each Convolution layer with implicit-GEMM performance metrics.

    Ported from trex/df_preprocessing.annotate_convolutions. For every layer of
    type ``Convolution`` this computes the number of fused multiply-accumulates
    (MACs), the arithmetic intensity (MACs/byte), the compute/memory efficiency
    (per millisecond) and the equivalent GEMM dimensions M, N, K.

    Returns a ``list`` of ``dict`` (one per convolution) carrying the layer name,
    precision, latency and footprint plus the computed ``attr.*`` fields. The
    list is empty if the engine has no convolutions.
    """
    latency_by_name = {rec["Name"]: rec for rec in plan.records}
    out = []
    for layer in plan.layers:
        if layer.type != "Convolution" or not layer.inputs or not layer.outputs:
            continue
        rec = latency_by_name.get(layer.name, {})
        N, C, H, W = layer.inputs[0].shape
        _, K, P, Q = layer.outputs[0].shape
        R, S = layer.raw_dict.get("Kernel", (1, 1))
        G = layer.raw_dict.get("Groups", 1)

        weights_vol = (K * C * R * S) / G
        input_vol = N * C * H * W
        output_vol = N * K * P * Q
        in_size = layer.inputs[0].data_size
        out_size = layer.outputs[0].data_size
        nb_bytes = input_vol * in_size + output_vol * out_size + weights_vol * in_size
        nb_macs = N * K * P * Q * C * R * S / G
        latency = float(rec.get("latency.avg_time", 0) or 0)

        out.append({
            "Name": layer.name,
            "precision": layer.precision,
            "latency.avg_time": latency,
            "latency.pct_time": float(rec.get("latency.pct_time", 0) or 0),
            "total_footprint_bytes": layer.total_footprint_bytes,
            "attr.groups": G,
            "attr.kernel": (R, S),
            "attr.macs": int(nb_macs),
            "attr.arithmetic_intensity": nb_macs / nb_bytes if nb_bytes else 0,
            "attr.compute_efficiency": nb_macs / latency if latency > 0 else 0,
            "attr.memory_efficiency": nb_bytes / latency if latency > 0 else 0,
            # Convolution expressed as a matrix multiply (M, K) x (K, N).
            "attr.M": int(N * P * Q),
            "attr.N": int(K),
            "attr.K": int(C * R * S),
        })
    return out

# ------------------------------------------------------------------------------
# Layer linters (performance-hazard heuristics)
# ------------------------------------------------------------------------------
#
# Ported from trex/lint.py (ConvLinter / ReformatLinter / SliceLinter / QDQLinter),
# rewritten to operate on EnginePlan records instead of a pandas DataFrame. Each
# linter returns a ``list`` of hazard ``dict`` (name / hazard / mitigation / help
# plus hazard-specific fields); an empty list means no hazards were found.

# Tactic-name substrings that indicate Tensor Core acceleration.
_TENSOR_CORE_TACTICS = ("imma", "hmma", "xmma", "i88", "884")

def lint_convolutions(plan: "EnginePlan") -> List[Dict]:
    """Flag convolutions that may under-perform (Tensor Core / precision / alignment)."""
    report = {}

    def is_small_conv(inputs):
        if len(inputs[0].shape) != 4:
            return False
        _, c, _, _ = inputs[0].shape
        return c < 32

    for conv in plan.get_layers_by_type("Convolution"):
        inputs, outputs = create_activations(conv)
        tactic = str(conv.get("tactic", ""))
        precision = conv.get("precision")

        # 1. Not accelerated by Tensor Cores (non-FP32 conv without a TC tactic).
        if precision != "FP32" and not any(t in tactic for t in _TENSOR_CORE_TACTICS):
            mitigation = ""
            if is_small_conv(inputs):
                mitigation = "This Convolution has a small number of input channels so acceleration may not be possible."
            report[conv["Name"]] = {
                "name": conv["Name"],
                "tactic": tactic,
                "subtype": conv.get("subtype"),
                "hazard": "Convolution is not accelerated.",
                "mitigation": mitigation,
                "help": "TensorCores accelerate large Convolution and GEMM operations.",
            }

        # 2. Quantized convolution with float outputs.
        if precision == "INT8" and inputs and outputs:
            if inputs[0].format[:4] == "Int8" and outputs[0].format[:4] != "Int8":
                report[conv["Name"]] = {
                    "name": conv["Name"],
                    "tactic": tactic,
                    "subtype": conv.get("subtype"),
                    "hazard": "Quantized Convolution has float outputs.",
                    "mitigation": "Consider adding quantization after the convolution.",
                    "help": "Quantized Convolution with float outputs is ill advised for memory-limited convolutions.",
                }

        # 3. Channel alignment not optimal for Tensor Cores.
        if inputs and outputs and len(inputs[0].shape) == 4 and len(outputs[0].shape) == 4:
            _, C, _, _ = inputs[0].shape
            _, K, _, _ = outputs[0].shape
            aligned = (C % 16 == 0 and K % 16 == 0) if precision == "INT8" else (C % 8 == 0 and K % 8 == 0)
            if not aligned:
                report.setdefault(conv["Name"], {
                    "name": conv["Name"],
                    "tactic": tactic,
                    "subtype": conv.get("subtype"),
                    "hazard": "Convolution channels are not optimally aligned.",
                    "mitigation": "Consider changing the alignment of the convolution's channels.",
                    "help": "For best performance, input/output channels of a Tensor Core convolution "
                    "should be aligned to 8 (FP32/FP16) or 16 (INT8).",
                })
    return list(report.values())

def _lint_type_conversion(layers: List[Dict], hazard_layer: str, extra=None) -> List[Dict]:
    """Shared helper: flag layers whose input/output data types differ."""
    report = []
    for layer in layers:
        inputs, outputs = create_activations(layer)
        if not inputs or not outputs:
            continue
        inf, outf = inputs[0].format[:4], outputs[0].format[:4]
        if inf != outf:
            mitigation = "Consider adding quantization around float operations." if "INT8" in (inf, outf) else ""
            entry = {
                "name": layer["Name"],
                "type conversion": f"{inf} -> {outf}",
                "shape conversion": f"{inputs[0].shape} -> {outputs[0].shape}",
                "hazard": f"{hazard_layer} layer is converting operand data type.",
                "mitigation": mitigation,
                "help": "Conversions between float32 and float16 are a red flag, as are conversions "
                "between float32/16 and INT8.",
            }
            if extra:
                entry.update(extra(layer))
            report.append(entry)
    return report

def lint_reformats(plan: "EnginePlan") -> List[Dict]:
    """Flag Reformat layers that convert between data types (not just layouts)."""
    return _lint_type_conversion(
        plan.get_layers_by_type("Reformat"),
        "Reformat",
        extra=lambda l: {"origin": l.get("Origin", "")},
    )

def lint_slices(plan: "EnginePlan") -> List[Dict]:
    """Flag Slice layers that convert between data types."""
    return _lint_type_conversion(plan.get_layers_by_type("Slice"), "Slice")

def lint_qdq(plan: "EnginePlan") -> List[Dict]:
    """Flag dangling (unfused) Quantize / Dequantize (Scale) layers."""
    report = []
    for scale in plan.get_layers_by_type("Scale"):
        inputs, outputs = create_activations(scale)
        if not inputs or not outputs:
            continue
        in_int8 = "Int8" in inputs[0].format
        out_int8 = "Int8" in outputs[0].format
        if in_int8 ^ out_int8:  # exactly one side is INT8 -> a Q or DQ conversion
            role = "Dequantize" if in_int8 else "Quantize"
            report.append({
                "name": scale["Name"],
                "type conversion": f"{inputs[0].format[:4]} -> {outputs[0].format[:4]}",
                "hazard": f"Unfused {role} layer",
                "mitigation": f"Check why the {role} layer is not fused.",
                "help": "Unfused Quantize/Dequantize nodes are wasteful and should be avoided. "
                "Quantize nodes may be necessary for quantizing inputs.",
            })
    return report

def lint_engine(plan: "EnginePlan") -> Dict[str, List[Dict]]:
    """Run all layer linters, returning a dict of ``{linter_name: [hazards]}``."""
    return {
        "Convolution": lint_convolutions(plan),
        "Reformat": lint_reformats(plan),
        "Slice": lint_slices(plan),
        "QDQ": lint_qdq(plan),
    }

# ------------------------------------------------------------------------------
# trtexec log parsing (build / profiling metadata)
# ------------------------------------------------------------------------------
#
# Ported from trex/utils/parse_trtexec_log.py, reading plain log files (the
# original also supported reading from a .tea zip archive). trtexec logs are
# organized into "=== Section ===" blocks of "[ts] [I] key: value" lines; these
# parsers extract the sections into the metadata dictionaries that EnginePlan
# consumes (builder_cfg / performance_summary / device_properties).

def _to_float(line: str) -> float:
    """Extract the first float found in a string."""
    m = re.search(r"[-+]?(\d+(\.\d*)?|\.\d+)([eE][-+]?\d+)?", line)
    if m is None:
        raise ValueError(f"no float in {line!r}")
    start, end = m.span()
    return float(line[start:end])

def _get_stats(line: str) -> List[float]:
    """Parse a 'k = v, k = v, ...' line into the list of float values."""
    return [_to_float(substr.split("=")[1]) for substr in line.split(",")]

class _FileSection:
    """Accumulates the key/value lines that belong to one trtexec log section."""

    def __init__(self, section_header: str):
        self.section_header = section_header
        self.dict = {}

    def entered_section(self, line: str) -> bool:
        return re.search(self.section_header, line) is not None

    def parse_line(self, line: str) -> bool:
        match = re.search(r"(\[\d+/\d+/\d+-\d+:\d+:\d+\] \[I\] )", line)
        if match is None:
            return False
        kv_line = line[match.span()[1]:].strip()
        if not kv_line.count(":"):
            return False
        kv = kv_line.split(":")
        if len(kv) > 1:
            self.dict[kv[0]] = kv[1][1:]
            return True
        return True  # a section sub-header line (still inside the section)

def _parse_log_file(file_name: str, sections: List[_FileSection]):
    current = None
    with open(file_name, errors="ignore") as f:
        for line in f.readlines():
            if current is None:
                current = next((s for s in sections if s.entered_section(line)), None)
            elif not current.parse_line(line):
                sections.remove(current)
                current = next((s for s in sections if s.entered_section(line)), None)

def parse_build_log(file_name: str) -> Dict:
    """Parse a trtexec *build* log, returning model/build/device dictionaries."""
    model_options = _FileSection("=== Model Options ===")
    build_options = _FileSection("=== Build Options ===")
    device_information = _FileSection("=== Device Information ===")
    _parse_log_file(file_name, [model_options, build_options, device_information])
    return {
        "model_options": model_options.dict,
        "build_options": build_options.dict,
        "device_information": device_information.dict,
    }

def parse_profiling_log(file_name: str) -> Dict:
    """Parse a trtexec *profiling* log, returning perf/inference/device dictionaries."""
    performance_summary = _FileSection("=== Performance summary ===")
    inference_options = _FileSection("=== Inference Options ===")
    device_information = _FileSection("=== Device Information ===")
    _parse_log_file(file_name, [performance_summary, inference_options, device_information])

    perf = performance_summary.dict
    for k, v in perf.items():
        try:
            if k in ("Throughput", "Total Host Walltime", "Total GPU Compute Time"):
                perf[k] = _to_float(v)
            elif k in ("Latency", "Enqueue Time", "H2D Latency", "GPU Compute Time", "D2H Latency"):
                perf[k] = _get_stats(v)
        except (ValueError, IndexError):
            pass
    dev = device_information.dict
    for k, v in dev.items():
        if k in ("Compute Clock Rate", "Memory Bus Width", "Memory Clock Rate", "Compute Capability", "SMs"):
            try:
                dev[k] = _to_float(v)
            except ValueError:
                pass
    return {
        "performance_summary": perf,
        "inference_options": inference_options.dict,
        "device_information": dev,
    }

def write_build_metadata(build_log: str, out_file: str) -> Dict:
    """Parse a build log and write the metadata JSON that EnginePlan consumes."""
    meta = parse_build_log(build_log)
    with open(out_file, "w") as f:
        json.dump(meta, f, indent=2)
    return meta

def write_profiling_metadata(profiling_log: str, out_file: str) -> Dict:
    """Parse a profiling log and write the metadata JSON that EnginePlan consumes."""
    meta = parse_profiling_log(profiling_log)
    with open(out_file, "w") as f:
        json.dump(meta, f, indent=2)
    return meta

# ------------------------------------------------------------------------------
# ONNX export (for viewing an engine graph in Netron)
# ------------------------------------------------------------------------------
#
# A compact re-implementation of trex/graphing.py's OnnxGraph: emit one ONNX node
# per engine layer, wired together by tensor names, so the engine can be opened
# in Netron. onnx is imported lazily.

# Map short data-type labels to ONNX TensorProto dtypes.
_ONNX_TYPE_KEYWORDS = (
    ("int8", "INT8"),
    ("fp32", "FLOAT"),
    ("float", "FLOAT"),
    ("fp16", "FLOAT16"),
    ("half", "FLOAT16"),
    ("int64", "INT64"),
    ("int32", "INT32"),
    ("bf16", "BFLOAT16"),
)

# ONNX op types that Netron colors specially; map TensorRT layer types onto them.
_ONNX_OP_MAP = {"Convolution": "Conv", "MatrixMultiply": "MatMul"}

def _onnx_tensor(onnx, activation: "Activation"):
    desc = activation.format.lower()
    dtype_name = next((name for kw, name in _ONNX_TYPE_KEYWORDS if kw in desc), "UNDEFINED")
    dtype = getattr(onnx.TensorProto, dtype_name)
    return onnx.helper.make_tensor_value_info(activation.name, dtype, activation.shape)

def export_engine_to_onnx(plan: "EnginePlan", output_path: str, include_constants: bool = False) -> str:
    """Export an engine plan to an ONNX file for viewing in Netron.

    One ONNX node is created per engine layer, connected by tensor names; the
    engine's input/output bindings become the ONNX graph inputs/outputs. This is
    a visualization aid - the ONNX is not runnable. Returns the written path.
    """
    import onnx  # lazy

    layers = list(plan.layers)
    if include_constants:
        layers = layers + list(plan.constants)

    nodes = []
    for layer in layers:
        op_type = _ONNX_OP_MAP.get(layer.type, layer.type)
        if op_type == "Pooling":
            op_type = "AveragePool" if layer.raw_dict.get("PoolingType") == "AVERAGE" else "MaxPool"
        if op_type == "Constant":
            continue
        node = onnx.helper.make_node(
            op_type,
            [i.name for i in layer.inputs],
            [o.name for o in layer.outputs],
            name=layer.name,
        )
        # Attach the simple scalar/list attributes; skip nested structures.
        skip = {"InputRegions", "OutputRegions", "Inputs", "Outputs", "Name", "ParameterType", "LayerName"}
        for key, value in sorted(layer.raw_dict.items()):
            if key in skip:
                continue
            try:
                node.attribute.extend([onnx.helper.make_attribute(key, value)])
            except (ValueError, TypeError):
                pass  # un-encodable attribute (e.g. a nested dict) - ignore
        nodes.append(node)

    g_inputs, g_outputs = plan.get_bindings()
    graph_def = onnx.helper.make_graph(
        nodes,
        plan.name or "trt-engine",
        [_onnx_tensor(onnx, t) for t in g_inputs],
        [_onnx_tensor(onnx, t) for t in g_outputs],
    )
    model = onnx.helper.make_model(graph_def, producer_name="tensorrt-cookbook-trex")
    base = output_path if str(output_path).endswith(".onnx") else f"{output_path}.onnx"
    onnx.save(model, base)
    return base

# ------------------------------------------------------------------------------
# Excel summary export
# ------------------------------------------------------------------------------
#
# Ported from trex/excel_summary.py, rewritten with openpyxl instead of
# pandas.ExcelWriter + xlsxwriter, and embedding Matplotlib PNGs instead of
# plotly images. openpyxl is imported lazily.

# The per-layer scalar columns written to the "Layers" worksheet.
_EXCEL_LAYER_COLS = (
    "Name",
    "type",
    "subtype",
    "precision",
    "output_precision",
    "tactic",
    "latency.avg_time",
    "latency.pct_time",
    "total_io_size_bytes",
    "weights_size",
    "total_footprint_bytes",
)

def write_engine_excel(plan: "EnginePlan", path: str, image_files: Dict[str, str] = None, columns=None) -> str:
    """Write an Excel summary of an engine plan (Summary / Layers / Precision + images).

    ``image_files`` is an optional ``{sheet_name: png_path}`` dict of figures to
    embed (e.g. the Matplotlib PNGs produced by the report-card examples).
    Returns the written path.
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image

    columns = columns or _EXCEL_LAYER_COLS
    wb = Workbook()

    # Sheet 1: Summary (model / device / builder / performance key-value pairs).
    ws = wb.active
    ws.title = "Summary"
    row = 1
    sections = (
        ("Model", summary_dict(plan)),
        ("Device Properties", plan.device_properties),
        ("Builder Configuration", plan.builder_cfg),
        ("Performance Summary", plan.performance_summary),
    )
    for section, data in sections:
        ws.cell(row=row, column=1, value=section)
        row += 1
        for k, v in data.items():
            ws.cell(row=row, column=1, value=str(k))
            ws.cell(row=row, column=2, value=str(v))
            row += 1
        row += 1

    # Sheet 2: Layers (curated per-layer table).
    ws = wb.create_sheet("Layers")
    ws.append(list(columns))
    for rec in plan.records:
        ws.append([_excel_cell(rec.get(c)) for c in columns])

    # Sheet 3: Precision statistics.
    ws = wb.create_sheet("Precision")
    stats = compute_precision_stats(plan)
    ws.append(["category", "precision", "bytes"])
    for category, data in stats.items():
        for precision, size in data.items():
            ws.append([category, str(precision), int(size)])

    # Optional image sheets.
    if image_files:
        for sheet_name, image_path in image_files.items():
            ws = wb.create_sheet(sheet_name[:31])  # Excel sheet-name length limit
            ws.add_image(Image(image_path), "A1")

    base = path if str(path).endswith(".xlsx") else f"{path}.xlsx"
    wb.save(base)
    return base

def _excel_cell(value):
    """Coerce a record value into something openpyxl can store."""
    if isinstance(value, (str, int, float)) or value is None:
        return value
    return str(value)

# ------------------------------------------------------------------------------
# GPU device info and clock state (NVML)
# ------------------------------------------------------------------------------
#
# Ported from trex/utils/device_info.py and utils/config_gpu.py, using pynvml
# (read-only) instead of pycuda. These query the GPU used to *build/profile* an
# engine; they need a GPU + pynvml. Locking clocks (for reproducible profiling)
# additionally needs root privileges and is intentionally left out here - the
# read-only helpers below expose current/max clocks so you can check throttling.

def query_device_info() -> List[Dict]:
    """Return a metadata dict per visible GPU (name, memory, clocks, ...)."""
    import pynvml

    pynvml.nvmlInit()
    try:
        out = []
        for i in range(pynvml.nvmlDeviceGetCount()):
            h = pynvml.nvmlDeviceGetHandleByIndex(i)
            mem = pynvml.nvmlDeviceGetMemoryInfo(h)
            name = pynvml.nvmlDeviceGetName(h)
            out.append({
                "Name": name.decode() if isinstance(name, bytes) else name,
                "TotalMemory": mem.total,
                "MaxSMClockMHz": pynvml.nvmlDeviceGetMaxClockInfo(h, pynvml.NVML_CLOCK_SM),
                "MaxMemClockMHz": pynvml.nvmlDeviceGetMaxClockInfo(h, pynvml.NVML_CLOCK_MEM),
            })
        return out
    finally:
        pynvml.nvmlShutdown()

def sample_gpu_state(gpu_id: int = 0) -> Dict:
    """Sample the current GPU temperature / power / clocks / utilization."""
    import pynvml

    pynvml.nvmlInit()
    try:
        h = pynvml.nvmlDeviceGetHandleByIndex(gpu_id)
        return {
            "temperature_C": pynvml.nvmlDeviceGetTemperature(h, pynvml.NVML_TEMPERATURE_GPU),
            "power_W": pynvml.nvmlDeviceGetPowerUsage(h) // 1000,
            "gpu_util_pct": pynvml.nvmlDeviceGetUtilizationRates(h).gpu,
            "sm_clock_mhz": pynvml.nvmlDeviceGetClockInfo(h, pynvml.NVML_CLOCK_SM),
            "mem_clock_mhz": pynvml.nvmlDeviceGetClockInfo(h, pynvml.NVML_CLOCK_MEM),
            "graphics_clock_mhz": pynvml.nvmlDeviceGetClockInfo(h, pynvml.NVML_CLOCK_GRAPHICS),
        }
    finally:
        pynvml.nvmlShutdown()

def get_max_clocks(gpu_id: int = 0) -> Tuple[int, int]:
    """Return the ``(max_sm_clock_mhz, max_mem_clock_mhz)`` of a GPU."""
    import pynvml

    pynvml.nvmlInit()
    try:
        h = pynvml.nvmlDeviceGetHandleByIndex(gpu_id)
        return (
            pynvml.nvmlDeviceGetMaxClockInfo(h, pynvml.NVML_CLOCK_SM),
            pynvml.nvmlDeviceGetMaxClockInfo(h, pynvml.NVML_CLOCK_MEM),
        )
    finally:
        pynvml.nvmlShutdown()

# ------------------------------------------------------------------------------
# TensorRT Engine Archive (TEA)
# ------------------------------------------------------------------------------
#
# Ported from trex/archiving.py. A TEA is a ZIP file bundling an engine plan
# together with its analysis artifacts (graph/profile JSON, build log, a
# plan-info JSON, ...) so a whole exploration session travels as one file.
# The archive itself is pure-zip; only plan introspection lazy-imports tensorrt.

class EngineArchive:
    """A ZIP-based TensorRT Engine Archive (``.tea``) reader/writer."""

    __version__ = "1.0"

    def __init__(self, archive_filename: str, mode: str = "w"):
        """``mode='w'`` creates a fresh archive; ``mode='r'`` opens one for reading."""
        assert mode in ("w", "r"), "mode must be 'w' (write) or 'r' (read)"
        self.archive_filename = archive_filename
        self.mode = mode
        self.zipf = None

    def open(self):
        import os
        from zipfile import ZIP_DEFLATED, ZipFile

        if self.mode == "w":
            if os.path.exists(self.archive_filename):
                os.remove(self.archive_filename)  # start a fresh archive
            self.zipf = ZipFile(self.archive_filename, "a", compression=ZIP_DEFLATED, compresslevel=9)
        else:
            self.zipf = ZipFile(self.archive_filename, "r")
        return self

    def close(self):
        assert self.zipf is not None
        self.zipf.close()
        self.zipf = None

    def __enter__(self):
        return self.open()

    def __exit__(self, *args):
        self.close()

    # -- writing / reading -----------------------------------------------------

    def _basename(self, fname: str) -> str:
        import os

        return os.path.basename(fname)

    def writef_txt(self, fname: str, text: str):
        """Write a text entry into the archive."""
        assert self.zipf is not None
        self.zipf.writestr(self._basename(fname), text)

    def writef_bin(self, fname: str, content: bytes):
        """Write a binary entry into the archive."""
        assert self.zipf is not None
        self.zipf.writestr(self._basename(fname), content)

    def add_file(self, path: str, arcname: str = None):
        """Copy an existing file on disk into the archive."""
        assert self.zipf is not None
        self.zipf.write(path, arcname or self._basename(path))

    def readf(self, fname: str) -> bytes:
        """Read an entry's bytes from the archive."""
        assert self.zipf is not None
        return self.zipf.read(self._basename(fname))

    def namelist(self) -> List[str]:
        """List the entries in the archive."""
        assert self.zipf is not None
        return self.zipf.namelist()

    # -- engine introspection --------------------------------------------------

    def archive_plan_info(self, plan_bytes: bytes, fname: str = "plan_cfg.json") -> Dict:
        """Deserialize an engine plan and archive its properties + IO tensors as JSON."""
        import tensorrt as trt

        logger = trt.Logger(trt.Logger.ERROR)
        with trt.Runtime(logger) as runtime:
            engine = runtime.deserialize_cuda_engine(plan_bytes)
            assert engine is not None, "Failed to deserialize the engine plan"
            # These weight-streaming properties raise unless the engine was built
            # with kWEIGHT_STREAMING, so skip them to avoid noisy TRT error logs.
            bad = {
                "weight_streaming_budget",
                "weight_streaming_budget_v2",
                "minimum_weight_streaming_budget",
                "streamable_weights_size",
                "weight_streaming_scratch_memory_size",
            }
            plan_dict = {}
            for attr in dir(engine):
                if attr in bad or attr.startswith("__"):
                    continue
                try:
                    if not callable(getattr(engine, attr)):
                        plan_dict[attr] = str(getattr(engine, attr))
                except Exception:
                    pass
            io_tensors = plan_dict["io_tensors"] = {}
            for index in range(engine.num_io_tensors):
                name = engine.get_tensor_name(index)
                io_tensors[name] = {
                    "mode": str(engine.get_tensor_mode(name)),
                    "dtype": str(engine.get_tensor_dtype(name)),
                    "shape": str(engine.get_tensor_shape(name)),
                    "location": str(engine.get_tensor_location(name)),
                }
        self.writef_txt(fname, json.dumps(plan_dict, ensure_ascii=False, indent=4))
        return plan_dict

# ------------------------------------------------------------------------------
# Engine tactic summary (trex `summary` sub-command)
# ------------------------------------------------------------------------------

def summarize_engine_tactics(plan: "EnginePlan", group_tactics: bool = False, sort_key: str = "count") -> List[Dict]:
    """Summarize the engine by tactic: count and latency-percent per tactic.

    Ported from trex/misc.summarize_engine_dict. When ``group_tactics`` is True,
    the per-tactic hash suffix (``_0x...``) is stripped so identical kernels with
    different tactic hashes are merged, and the *total* latency-% is reported;
    otherwise the *mean* latency-% per (unique) tactic is reported. ``sort_key``
    is one of ``count`` / ``latency`` / ``id``.

    Returns a ``list`` of ``{"tactic", "count", "latency %"}`` dicts.
    """
    records = plan.records
    if group_tactics:
        stripped = []
        for rec in records:
            tactic = str(rec.get("tactic", "TensorRT"))
            idx = tactic.find("_0x")
            stripped.append({**rec, "tactic": tactic[:idx] if idx >= 0 else tactic})
        records = stripped
        latency = group_sum(records, "tactic", "latency.pct_time")
        if sort_key == "id":
            sort_key = "count"
    else:
        latency = group_mean(records, "tactic", "latency.pct_time")
        if sort_key == "count":
            sort_key = "latency"

    counts = group_count(records, "tactic")
    rows = [{"tactic": t, "count": counts[t], "latency %": latency[t]} for t in latency]

    if sort_key == "latency":
        rows.sort(key=lambda r: r["latency %"], reverse=True)
    elif sort_key == "count":
        rows.sort(key=lambda r: r["count"], reverse=True)
    return rows
