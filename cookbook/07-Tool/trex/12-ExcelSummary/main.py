# Copyright (c) 2024 NVIDIA CORPORATION & AFFILIATES.
# All rights reserved.
#
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Export an engine plan to an Excel workbook (summary + per-layer table +
# precision stats + embedded figures).
#
# This is the cookbook re-implementation of `trt-engine-explorer`'s
# excel_summary.ExcelSummary, rewritten with openpyxl (instead of
# pandas.ExcelWriter + xlsxwriter) and embedding Matplotlib PNGs (instead of
# plotly images). Requires the `openpyxl` package (`pip install openpyxl`).

from pathlib import Path

from matplotlib import pyplot as plt

from tensorrt_cookbook import EnginePlan, case_mark, colors_for, group_sum, layer_colormap, write_engine_excel

data_path = Path(__file__).parent.parent / "data"  # shared engine JSON files
out_path = Path(__file__).parent  # this example's own output
graph_json = data_path / "model.graph.json"
profile_json = data_path / "model.profile.json"
xlsx_file = out_path / "engine_summary.xlsx"

def _load():
    return EnginePlan(str(graph_json), str(profile_json), name="model")

def _latency_figure(plan, png_path):
    """Draw a small latency-by-type figure to embed in the workbook."""
    lat = group_sum(plan.records, "type", "latency.avg_time")
    types = sorted(lat, key=lat.get)
    fig, ax = plt.subplots(figsize=(6, 4))
    ax.barh(types, [lat[t] for t in types], color=colors_for(types, layer_colormap))
    ax.set_xlabel("Latency (ms)")
    ax.set_title("Latency by layer type")
    fig.tight_layout()
    fig.savefig(png_path)
    plt.close(fig)

@case_mark
def case_excel_summary():
    """Write an Excel workbook summarising the engine, with an embedded figure."""
    plan = _load()

    # Draw a figure to embed on its own worksheet.
    png_path = out_path / "latency_by_type.png"
    _latency_figure(plan, png_path)

    out = write_engine_excel(
        plan,
        str(xlsx_file),
        image_files={"Latency chart": str(png_path)},
    )
    print(f"Wrote {out}")

    # Report what was written.
    import openpyxl
    wb = openpyxl.load_workbook(out)
    print(f"Worksheets: {wb.sheetnames}")
    print(f"Layers sheet: {wb['Layers'].max_row - 1} layers x {wb['Layers'].max_column} columns")

if __name__ == "__main__":
    case_excel_summary()

    print("Finish")
